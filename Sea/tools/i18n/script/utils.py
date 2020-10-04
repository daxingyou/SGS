#!/usr/bin/env python
# coding=utf-8

import os
import zipfile
import shutil
import hashlib
import platform
import re
import StringIO
import collections
import ConfigParser
import subprocess
import sys
import getopt
import uuid
import json
import urllib
import base64
import csv
import codecs
from xml.etree import ElementTree as ET
from xml.etree.ElementTree import SubElement
from xml.etree.ElementTree import Element
from xml.etree.ElementTree import ElementTree
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def checkCode(msg):
    if platform.system() == "Windows":
        return msg.encode('gbk','ignore')
    else:
        return msg

def getSVNInfo():
    info = urllib.urlopen('http://10.235.102.201/client')
    return json.loads(base64.b64decode(info.read()))

def printSplit(name):
    length = len(name)
    length = int((100 - length))
    print(checkCode("\033[31m%s %s\033[34m" % (name, "-"*length)))

#  打印文件


def printFile(file, name):
    length = len(name)
    length = int((100 - length) / 2)
    print(checkCode("\033[35m%s%s%s\033[0m" % (">"*length, name, "<"*length)))
    f = open(file, 'r')
    print f.read()
    f.close()
    print(checkCode("\033[35m%s%s%s\033[0m" % (">"*length, name, "<"*length)))
#  打印对象


def printObject(object, name):
    print(checkCode("%s = %s" % (name, json.dumps(object, ensure_ascii=False, indent=4))))


def error(msg):
    print(checkCode("error: %s " % msg))
    exit(-1)


def info(msg):
    print(checkCode("info: %s " % msg))


def initUtf8():
    import sys
    reload(sys)
    sys.setdefaultencoding('utf8')


def byteify(input):
    if isinstance(input, dict):
        return {byteify(key): byteify(value) for key, value in input.iteritems()}
    elif isinstance(input, list):
        return [byteify(element) for element in input]
    elif isinstance(input, unicode):
        return input.encode('utf-8')
    else:
        return input

def getOptions():
    optConfig = ""
    optPath = ""
    optOut = ""
    options,args = getopt.getopt(sys.argv[1:], "c:p:o:", ["config=","path=","out="])
    for name, value in options:
        if name in ("-c","--config"):
            optConfig = value
        if name in ("-p","--path"):
            optPath = os.path.realpath(value)
        if name in ("-o","--out"):
            optOut = os.path.realpath(value)
    print("option: optConfig=%s, optPath=%s, optOut=%s" % (optConfig, optPath, optOut)) 
    return optConfig, optPath, optOut

def getConfigJson(DIR, cfgName):
    bname = os.path.basename(cfgName)
    iniName = os.path.splitext(bname)[0] + ".json"
    cfgPath = os.path.join(DIR, iniName)
    cfg = json.load(open(cfgPath, 'r'))
    cfg = byteify(cfg)
    return cfg, cfgName, cfgPath


def punctuationChinaToEnglish(cnStr,isNewLine=True):

    #   $str = ~ s /【/[/g;
    # $str = ~ s /】/]/g;
    # $str =~ s/。/./g;
    # $str =~ s/？/?/g;
    # $str =~ s/！/!/g;
    # $str =~ s/（/(/g;
    # $str =~ s/）/)/g;
    # $str =~ s/、/, /g;
    # $str =~ s/\*/x/g;
    # # 标点符号后加一个空格
    # $str =~ s/：/: /g;
    # $str =~ s/，/, /g;
    # # 删除标点符号
    # $str =~ s/“//g;
    # $str =~ s/”//g;
    # $str =~ s/《//g;
    # $str =~ s/》//g;
    # table = {ord(f): ord(t) for f, t in zip(
    #     u'，。！？【】（）％＃＠＆１２３４５６７８９０',
    #     u',.!?[]()%#@&1234567890')}
    
    table = {ord(f): ord(t) for f, t in zip(
        u'。！？【】（）*：',
        u'.!?[]()x:')}

    print(checkCode("pun cn to en  cnstr : %s" % cnStr))
    cnStr = unicode(cnStr)
    enStr = cnStr.translate(table)
    # 标点符号后加一个空格
    enStr = enStr.replace('、', ', ')
    enStr = enStr.replace('，', ', ')
    # 删除标点符号
    enStr = enStr.replace('“', '')
    enStr = enStr.replace('”', '')
    enStr = enStr.replace('《', '')
    enStr = enStr.replace('》', '')
    #特殊处理
    if isNewLine:
        enStr = enStr.replace('\\n', '\n')
        enStr = enStr.replace('\\\"', '\"')
    print(checkCode("pun cn to en enstr : %s" % enStr))

    # return enStr.encode('UTF-8')
    return enStr

# 是否含有数字
def isMathNumber(str):
    return re.compile('\d').search(str)

# 是否含有字母
def isMathLetter(str):
    return re.compile('[a-zA-Z]').search(str)

# 是否含中文
def isMathChinese(str):
    # print("isMathChinese  str = %s" % str)
    return not re.compile(u'[\u4e00-\u9fa5]').search(str.decode('UTF-8')) == None
    # return not len(str) == len(str.decode('UTF-8'))

def isChinese(str):
    return not len(str) == len(str.decode('UTF-8'))

def isMathChinese2(str):
    str = str.decode("utf-8")
    for ch in str:
        if u'\u4e00' <= ch <= u'\u9fa5':
            return True
    return False

# 判断是否是全数字：
def isAllNumber(str):
    return str.encode('UTF-8').isdigit()

#  判断是否是全英文：
def isAllLetter(str):
    return str.encode('UTF-8').isalpha()

#  判断是否是全中文：
def isAllChinese(str):
    is_MathChinese = isMathChinese(str)
    return is_MathChinese and len(str) == len(str.decode('UTF-8'))

def writeJsonFile(fileData, fullPath):
    print("writeJsonFile = %s" % fullPath)
    f = open(fullPath, "w")
    with open(fullPath, "w") as f:
        json.dump(fileData, f, ensure_ascii=False, indent=4)
    f.close()


def readJsonFile(DIR, name, ext=None):
    jsonPath = getFilesWithName(DIR, name, ext)[0]
    print("readJsonFile = %s" % jsonPath)
    data = json.load(open(jsonPath, 'r'))
    data = byteify(data)
    return data, name, jsonPath


def readCsvFile(DIR, name, ext=None):
    # isinstance(input, list):
    print("readCsvFile  DIR = %s" % DIR)
    print("readCsvFile  name = %s" % name)
    printObject(getFilesWithName(DIR, name, ext), "readCsvFile")
    csvPath = getFilesWithName(DIR, name, ext)[0]
    print("readCsvFile = %s" % csvPath)
    ret = []
    # with open(csvPath) as f:
    with open(csvPath, "rU") as f:
        reader = csv.reader(f)
        # 读取一行，下面的reader中已经没有该行了
        head_row = next(reader)
        print(head_row)
        for row in reader:
            # 行号从2开始
            # printObject(row, reader.line_num)
            ret.append(row)
    return ret, head_row


def writeCsvFile(datas, fullPath, headers):
    print("writeCvsFile = %s" % fullPath)
    file = open(fullPath, 'wb+')
    file.write(codecs.BOM_UTF8)
    with file as f:
        writer = csv.writer(f, dialect='excel')
        if headers:
            # writer = csv.DictWriter(f, headers)
            # writer.writeheader()
            writer.writerow(headers)
        for row in datas:
            writer.writerow(row)
        # 还可以写入多行
        # writer.writerows(datas)
    file.close()


def writeXmlFile(xmlTree, fullPath):
    print("writeXmlFile = %s" % fullPath)
    if xmlTree:
       xmlTree.write(fullPath)
    # xmlTree.close()


def readXmlFile(DIR, name, ext=None):
    xmlPath = getFilesWithName(DIR, name, ext)[0]
    print("readXmlFile = %s" % xmlPath)
    tree = ET.parse(xmlPath)
    return tree


def writeXlsxFile(xlsxTree, fullPath, headers):
    print("writeXlsxFile = %s" % fullPath)
    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = Workbook()
    #获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    #可以使用写入headers
    ws.append(headers)
    for row in xlsxTree:
        ws.append(row)
    # rows = len(xlsxTree)
    # for idx in range(0, rows):
    #     ws.append(xlsxTree[idx])
    #保存
    wb.save(filename=fullPath)

def readXlsxFile(DIR, name, ext=None):
    xlsxFileList = getFilesWithName(DIR, name, ext)
    info("readXlsxFile: " + str(xlsxFileList))
    if len(xlsxFileList) > 0:
        xlsxPath = xlsxFileList[0]
    else:
        error(DIR+"中"+ name + "不存在")
        return    
    print("readXlsxFile = %s" % xlsxPath)
    #打开一个workbook
    wb = load_workbook(filename=xlsxPath)
    #获取当前活跃的worksheet,默认就是第一个worksheet
    #ws = wb.active
    encoding = wb.encoding
    #当然也可以使用下面的方法
    #获取所有表格(worksheet)的名字
    sheets = wb.get_sheet_names()

    sheet_len = len(sheets)
    sheet_list = []
    for idx in range(0, sheet_len):
        #表格的名称
        sheet_first = sheets[idx]
        #获取特定的worksheet
        ws = wb.get_sheet_by_name(sheet_first)
        #获取表格所有行和列，两者都是可迭代的
        rows = ws.rows
        columns = ws.columns
        ret = []
        # print rows
        #迭代所有的行
        for row in rows:
            # print row
            # line = [col.value for col in row]
            line = [col.value for col in row]
            ret.append(byteify(line))

        sheet_list.append(ret)
    #如果只有一个表 直接返回第一个表数据 
    if sheet_len == 1:
        sheet_list = sheet_list[0]
    return sheet_list

def readXlsxFile2(xlsxPath,dataOnly=False):
    print("readXlsxFile2 xlsxPath = %s" % xlsxPath)
    #打开一个workbook
    wb = load_workbook(filename=xlsxPath,data_only=dataOnly)
    #获取当前活跃的worksheet,默认就是第一个worksheet
    #ws = wb.active
    encoding = wb.encoding
    #当然也可以使用下面的方法
    #获取所有表格(worksheet)的名字
    sheets = wb.get_sheet_names()

    sheet_len = len(sheets)
    sheet_list = []
    for idx in range(0, sheet_len):
        #表格的名称
        sheet_first = sheets[idx]
        #获取特定的worksheet
        ws = wb.get_sheet_by_name(sheet_first)
        #获取表格所有行和列，两者都是可迭代的
        rows = ws.rows
        columns = ws.columns
        ret = []
        # print rows
        #迭代所有的行
        for row in rows:
            # print row
            # line = [col.value for col in row]
            line = [col.value for col in row]
            
            ret.append(byteify(line))

        sheet_list.append(ret)
    # #如果只有一个表 直接返回第一个表数据 
    # if sheet_len == 1:
    #     sheet_list = sheet_list[0]
    return sheet_list

    # #第一个表格的名称
    # sheet_first = sheets[0]
    # #获取特定的worksheet
    # ws = wb.get_sheet_by_name(sheet_first)

    # #获取表格所有行和列，两者都是可迭代的
    # rows = ws.rows
    # columns = ws.columns
    # ret = []

    # #迭代所有的行
    # for row in rows:
    #     # print row
    #     # line = [col.value for col in row]
    #     line = [ col.value for col in row]
    #     ret.append(line)

    # return ret


def getMd5(str):
    # print("getMd5 str = %s" % str)
    md = hashlib.md5()
    md.update(str)
    md5 = md.hexdigest()   
    return md5


def writeFile(lines, fullPath):
    print("writeFile = %s" % fullPath)
    filename = fullPath
    lines = lines
    f = open(filename, "w")
    for line in lines:
        f.write(line)
    f.close()

def downloadConfig(url, path):
    print("downloadConfig: %s->%s" % (url, path))
    urllib.urlretrieve(url, path)


def getSubDir(ROOT, path):
    if ROOT == path:
        return ""
    return path[len(ROOT) + 1:]


def copyFile(oldname,newname):
    if os.path.exists(oldname):
        shutil.copyfile(oldname,newname)

def removeDir(DIR):
    if os.path.exists(DIR):
        shutil.rmtree(DIR)


def copyDir(olddir, newdir):
    print("olddir = %s" % olddir)
    print("newdir = %s" % newdir)
    if os.path.exists(newdir):
        shutil.rmtree(newdir)
    shutil.copytree(olddir, newdir)


def copyDir2(olddir, newdir):
    for dirpath, dirnames, filenames in os.walk(olddir):
        for filename in filenames:
            fullPath = os.path.join(dirpath, filename)
            if not os.path.isdir(fullPath):
                tempPath = os.path.abspath(os.path.join(
                    newdir, os.path.relpath(fullPath, olddir)))
                tempDir = os.path.dirname(tempPath)
                if not os.path.exists(tempDir):
                    os.makedirs(tempDir)

                if os.path.exists(fullPath):
                    shutil.copy(fullPath, tempPath)


def cleanDir(DIR, ignores):
    print("cleanDir = %s" % DIR)
    lists = os.walk(DIR)
    for root, dirs, files in lists:
        for f in files:
            path = os.path.join(root, f)
            tmp = getSubDir(DIR, path)
            finded = False
            for v in ignores:
                if tmp.startswith(v):
                    finded = True
                    break
            if finded:
                continue
            os.remove(path)
        for d in dirs:
            path = os.path.join(root, d)
            tmp = getSubDir(DIR, path)
            finded = False
            for v in ignores:
                if tmp.startswith(v):
                    finded = True
                    break
            if finded:
                continue
            shutil.rmtree(path)


def removeFile(full_path):
    print("removeFile = %s" % full_path)
    if os.path.isfile(full_path):
        os.remove(full_path)


def checkFileByExtention(ext, path):
    filelist = os.listdir(path)
    for fullname in filelist:
        name, extention = os.path.splitext(fullname)
        if extention == ext:
            return name, fullname
    return (None, None)

def modifyFile1(filename, func):
    f = open(filename, "r")
    lines = f.readlines()
    f.close()
    f = open(filename, "w")
    for line in lines:
        func(f, line)
    f.close()


def restoreFile(backUp):
    for v in backUp:
        filename = v["filename"]
        lines = v["lines"]
        f = open(filename, "w")
        for line in lines:
            f.write(line)
        f.close()

def getExecPath():
    import sys
    if hasattr(sys, "frozen"):
        return os.path.realpath(sys.executable)
    else:
        return os.path.realpath(sys.argv[0])


def getFilesWithName(DIR, name, ext=None):
    # filelist = os.listdir(DIR)
    if name == None and ext == None:
        return DIR
    if ext == None:
        _list = name.split(".")
        ext = "." + _list[-1]
        name = name[0:(len(name)-len(ext))]
    ret = []
    # print("getFilesWithName  name = %s" % name)
    # print("getFilesWithName  ext = %s" % ext)
    for dirpath, dirnames, filenames in os.walk(DIR):
        for filename in filenames:
            if os.path.splitext(filename)[0] == name and os.path.splitext(filename)[1] == ext:
                # return os.path.join(dirpath, filename)
                ret.append(os.path.join(dirpath, filename))
    return ret


def _mathFilePath(file_path, paths):
    isMath = True
    if paths and file_path and len(file_path) >= len(paths):
        # printObject(file_path, "_mathFilePath file_path")
        # printObject(paths, "_mathFilePath paths")
        index = -1
        _len = len(paths)
        i = 0
        while i < _len:
            i = i + 1
            path_dir = paths[index]
            index = index - 1
            file_dir = file_path[index]
            # printObject(path_dir, "_mathFilePath path_dir")
            # printObject(file_dir, "_mathFilePath file_dir")
            if path_dir != file_dir:
                return False
    else:
        isMath = False

    return isMath


def seachPathBylist(path_list, name, path):
    # printObject(path_list, "seachPathBylist path_list")
    # printObject(name, "seachPathBylist name")
    # printObject(path, "seachPathBylist path")

    length = len(path_list)
    if length < 1:
        print('utils seachPathBylist path_list is a None ...')
        return None
    # if path == None and length == 1:
    #     return path_list[0]

    for full_path in path_list:
        file_path = full_path.split(os.path.sep)
        file_name = file_path[-1]
        if file_name == name and path == None:
            return full_path
        if path != None and path != "":
            paths = path.split(os.path.sep)
            # printObject(file_path, "seachPathBylist file_path")
            # printObject(paths, "seachPathBylist paths")
            is_math = _mathFilePath(file_path, paths)
            if is_math == True:
                return full_path

    return None


def removeFileWithExt(DIR, ext):
    file_list = os.listdir(DIR)
    for f in file_list:
        full_path = os.path.join(DIR, f)
        if os.path.isdir(full_path):
            removeFileWithExt(full_path, ext)
        elif os.path.isfile(full_path):
            name, cur_ext = os.path.splitext(f)
            if cur_ext == ext:
                os.remove(full_path)


# def getFilesWithExt(path, ext):
#     ret = []
#     for dirpath, dirnames, filenames in os.walk(path):
#         for filename in filenames:
#             if os.path.splitext(filename)[1] == ext:
#                 ret.append(os.path.join(dirpath, filename))

#     printObject(ret,"getFilesWithExt")
#     return ret

def getFilesWithExt(DIR, ext, name=None):
    # filelist = os.listdir(DIR)
    if name == None and ext == None:
        return DIR
    if name == None:
        name = ""
    ret = []
    # print("getFilesWithExt  name = %s" % name)
    # print("getFilesWithExt  ext = %s" % ext)
    for dirpath, dirnames, filenames in os.walk(DIR):
        for filename in filenames:
            # print("getFilesWithExt  filename = %s" % filename)
            # print("getFilesWithExt  ext = %s" % os.path.splitext(filename)[1])
            if os.path.splitext(filename)[1] == ext:
                filePath = os.path.join(dirpath, filename)
                # print("getFilesWithExt  filePath = %s" % filePath)
                if os.path.splitext(filename)[0] == name :
                    return filePath
                else:
                    ret.append(filePath)
    # printObject(ret,"getFilesWithExt:")                
    return ret



def mkOutDir(DIR_OUT, CLEANUP_FIRST, ignores=[]):
    if os.path.exists(DIR_OUT):
        if CLEANUP_FIRST:
            cleanDir(DIR_OUT, ignores)
            # shutil.rmtree(DIR_OUT)
            # os.makedirs(DIR_OUT)
    else:
        os.makedirs(DIR_OUT)

def formatName(s):
    l = s.split("_")
    ret = ""
    for v in l:
        ret += v[0].upper() + v[1:].lower()
    return ret

def runMain(main, waitWindows=True):
    initUtf8()
    try:
        main()
    except Exception as e:
        print(e)
        import traceback
        traceback.print_exc()
    finally:
        if waitWindows and platform.system() == "Windows":
            print('Press any key to continue...')
            raw_input()


def runShell(cmd, cwd=None, wait=False):
    p = subprocess.Popen(cmd, shell=True, cwd=cwd)
    if wait:
        p.wait()
        if p.returncode:
            raise subprocess.CalledProcessError(
                returncode=p.returncode,
                cmd=cmd)
        return p.returncode
    return 0
